from __future__ import annotations

from copy import copy
from pathlib import Path


def combine_vertical(src: Path, trans: Path, out: Path) -> Path:
    """Append translated sheets into src workbook with '_translated' suffix."""
    from openpyxl import load_workbook

    master = load_workbook(src)
    trans_wb = load_workbook(trans)

    for ws in trans_wb.worksheets:
        new_title = f"{ws.title}_translated"[:31]  # Excel sheet title limit
        new_ws = master.create_sheet(title=new_title)
        # Copy cells + styles
        for row in ws.iter_rows():
            for cell in row:
                new_cell = new_ws[cell.coordinate]
                new_cell.value = cell.value
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.fill = copy(cell.fill)
                    new_cell.border = copy(cell.border)
                    new_cell.alignment = copy(cell.alignment)
                    new_cell.number_format = cell.number_format
                    new_cell.protection = copy(cell.protection)
        # Copy column widths
        for letter, dim in ws.column_dimensions.items():
            new_ws.column_dimensions[letter].width = dim.width
        # Copy RTL setting
        new_ws.sheet_view.rightToLeft = ws.sheet_view.rightToLeft
    master.save(out)
    master.close()
    trans_wb.close()
    return out


def combine_horizontal(src: Path, trans: Path, out: Path, rtl: bool = False) -> Path:
    """Side-by-side per-row layout: original columns, then translated columns.

    Creates a new workbook where each sheet has original cells on the left and
    translated cells offset by (source_max_col + 1) to the right (LTR) or swapped (RTL).
    """
    from copy import copy as deep_copy_style
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    src_wb = load_workbook(src)
    trans_wb = load_workbook(trans)
    out_wb = Workbook()
    # Remove default sheet
    default = out_wb.active
    out_wb.remove(default)

    for src_ws in src_wb.worksheets:
        name = src_ws.title[:31]
        out_ws = out_wb.create_sheet(title=name)
        # Find matching translated sheet (same title).
        if name in trans_wb.sheetnames:
            trans_ws = trans_wb[name]
        else:
            trans_ws = trans_wb.worksheets[0]

        src_max_col = src_ws.max_column or 1
        trans_max_col = trans_ws.max_column or 1
        offset = src_max_col + 1  # leave one blank column separator

        if rtl:
            _copy_block(trans_ws, out_ws, col_offset=0)
            _copy_block(src_ws, out_ws, col_offset=trans_max_col + 1)
            out_ws.sheet_view.rightToLeft = True
        else:
            _copy_block(src_ws, out_ws, col_offset=0)
            _copy_block(trans_ws, out_ws, col_offset=offset)

        # Separator column header
        sep_col = (trans_max_col if rtl else src_max_col) + 1
        out_ws.cell(row=1, column=sep_col).value = ""

        # Header labels
        out_ws.cell(row=1, column=1).font = Font(bold=True)

    out_wb.save(out)
    out_wb.close()
    src_wb.close()
    trans_wb.close()
    return out


def _copy_block(src_ws, dst_ws, col_offset: int) -> None:
    from copy import copy as deep_copy_style

    for row in src_ws.iter_rows():
        for cell in row:
            dst_cell = dst_ws.cell(row=cell.row, column=cell.column + col_offset)
            dst_cell.value = cell.value
            if cell.has_style:
                dst_cell.font = deep_copy_style(cell.font)
                dst_cell.fill = deep_copy_style(cell.fill)
                dst_cell.border = deep_copy_style(cell.border)
                dst_cell.alignment = deep_copy_style(cell.alignment)
                dst_cell.number_format = cell.number_format
                dst_cell.protection = deep_copy_style(cell.protection)
    # Copy column widths (shifted)
    from openpyxl.utils import get_column_letter
    for letter, dim in src_ws.column_dimensions.items():
        if dim.width is None:
            continue
        try:
            from openpyxl.utils import column_index_from_string
            idx = column_index_from_string(letter)
            new_letter = get_column_letter(idx + col_offset)
            dst_ws.column_dimensions[new_letter].width = dim.width
        except Exception:
            continue
