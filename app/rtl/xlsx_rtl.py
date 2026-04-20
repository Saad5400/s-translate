from __future__ import annotations

from pathlib import Path


def apply(out_path: Path) -> Path:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment

    wb = load_workbook(out_path)
    for ws in wb.worksheets:
        ws.sheet_view.rightToLeft = True
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip():
                    # Preserve existing vertical alignment + other fields.
                    existing = cell.alignment
                    cell.alignment = Alignment(
                        horizontal="right",
                        vertical=existing.vertical,
                        wrap_text=existing.wrap_text,
                        shrink_to_fit=existing.shrink_to_fit,
                        indent=existing.indent,
                        text_rotation=existing.text_rotation,
                        readingOrder=2,
                    )
    wb.save(out_path)
    wb.close()
    return out_path
