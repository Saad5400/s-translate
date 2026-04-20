from __future__ import annotations

from pathlib import Path

from ..schemas import DocFormat, Segment
from .base import Translator


class XlsxTranslator(Translator):
    fmt = DocFormat.XLSX

    def extract(self, src_path: Path) -> list[Segment]:
        from openpyxl import load_workbook

        wb = load_workbook(src_path, data_only=False)
        segments: list[Segment] = []
        seg_idx = 0
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if not isinstance(val, str):
                        continue
                    if not val.strip():
                        continue
                    # Skip formulas (string starting with '=').
                    if val.startswith("="):
                        continue
                    segments.append(
                        Segment(
                            id=f"s{seg_idx}",
                            text=val,
                            meta={"sheet": ws.title, "coord": cell.coordinate},
                        )
                    )
                    seg_idx += 1
        wb.close()
        return segments

    def reinsert(self, src_path: Path, segments: list[Segment], out_path: Path) -> Path:
        from openpyxl import load_workbook

        # Build lookup: (sheet, coord) -> translated text
        lookup: dict[tuple[str, str], str] = {}
        for seg in segments:
            tr = seg.translated if seg.translated is not None else seg.text
            lookup[(seg.meta["sheet"], seg.meta["coord"])] = tr

        wb = load_workbook(src_path, data_only=False)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    key = (ws.title, cell.coordinate)
                    if key in lookup:
                        cell.value = lookup[key]
        wb.save(out_path)
        wb.close()
        return out_path
